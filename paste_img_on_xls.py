# -*- coding: utf-8 -*-
"""
Created on Sat Jan 12 11:06:15 2019

@author: hfuji
"""

import openpyxl
import glob

# https://qiita.com/itoru257/items/03bdd393becbe53d327f
'''
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
img = openpyxl.drawing.image.Image('model_plot.png')
ws.add_image( img, 'A1' )
wb.save('out.xlsx')
'''

# https://stackoverflow.com/questions/10888969/insert-image-in-openpyxl
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
#ws.merge_cells('A1:A3')

img_dir = 'D:/Develop/Tools/mAP/results/images/'
img_files = glob.glob(img_dir + '*.jpg')

for idx, img_fpath in enumerate(img_files):
    img = openpyxl.drawing.image.Image(img_fpath)
    row_number = idx*10 + 1
    col_idx = 1
    cell = ws.cell(row=row_number, column=col_idx)
    cell_str = 'A' + str(row_number)
    ws.add_image(img, cell_str)
    if idx >= 2:
        break
wb.save('output.xlsx')